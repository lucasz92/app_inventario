#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script para crear un archivo Excel de ejemplo con datos para importar
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def crear_excel_ejemplo():
    """Crea un archivo Excel de ejemplo con datos para importar."""
    
    # Crear workbook
    wb = Workbook()
    
    # Eliminar la hoja por defecto
    wb.remove(wb.active)
    
    # ======== HOJA DE PRODUCTOS ========
    ws_productos = wb.create_sheet(title="productos")
    
    # Encabezados de productos
    headers_productos = [
        "codigo", "descripcion", "categoria", "proveedor", "puesto_trabajo",
        "stock_actual", "unidad_medida", "tipo_control", "stock_minimo", 
        "stock_maximo", "observacion"
    ]
    
    # Agregar encabezados con estilo
    for col, header in enumerate(headers_productos, 1):
        cell = ws_productos.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2383E2", end_color="2383E2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de ejemplo para productos
    productos_ejemplo = [
        ["TORN001", "Tornillo Phillips M6x20", "Ferretería", "Ferretería Central", "Taller Mecánico", 500, "unidad", "manual", 50, 1000, "Tornillos estándar para uso general"],
        ["TURC001", "Tuerca hexagonal M6", "Ferretería", "Ferretería Central", "Taller Mecánico", 300, "unidad", "manual", 30, 600, "Tuercas galvanizadas"],
        ["CABL001", "Cable eléctrico 2.5mm", "Eléctrico", "ElectroSur", "Taller Eléctrico", 1000, "metro", "automático", 100, 2000, "Cable para instalaciones domiciliarias"],
        ["INTE001", "Interruptor simple 10A", "Eléctrico", "ElectroSur", "Taller Eléctrico", 75, "unidad", "manual", 10, 150, "Interruptores para uso doméstico"],
        ["PINT001", "Pintura látex blanca 4L", "Pintura", "Pinturas del Norte", "Almacén General", 25, "litro", "manual", 5, 50, "Pintura interior lavable"],
        ["RODI001", "Rodillo lana 23cm", "Pintura", "Pinturas del Norte", "Almacén General", 40, "unidad", "manual", 8, 80, "Rodillos para pintura látex"],
        ["TALA001", "Taladro percutor 800W", "Herramientas", "Herramientas Pro", "Taller Mecánico", 8, "unidad", "manual", 2, 15, "Taladros profesionales con percutor"],
        ["BROC001", "Broca HSS 8mm", "Herramientas", "Herramientas Pro", "Taller Mecánico", 50, "unidad", "manual", 10, 100, "Brocas para metal alta velocidad"],
        ["SOLD001", "Soldadura E6013 3.2mm", "Soldadura", "Soldaduras Técnicas", "Taller Soldadura", 200, "kilogramo", "automático", 20, 400, "Electrodos para soldadura general"],
        ["DISC001", "Disco corte metal 115mm", "Herramientas", "Herramientas Pro", "Taller Mecánico", 120, "unidad", "manual", 20, 250, "Discos para amoladora angular"],
        ["ACEI001", "Aceite hidráulico ISO 68", "Lubricantes", "Lubricantes Industriales", "Almacén Químicos", 80, "litro", "automático", 15, 150, "Aceite para sistemas hidráulicos"],
        ["FILT001", "Filtro de aire K&N", "Filtros", "Repuestos Automotor", "Taller Automotriz", 25, "unidad", "manual", 5, 50, "Filtros de aire de alto rendimiento"]
    ]
    
    # Agregar datos de productos
    for row, producto in enumerate(productos_ejemplo, 2):
        for col, valor in enumerate(producto, 1):
            ws_productos.cell(row=row, column=col, value=valor)
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers_productos) + 1):
        ws_productos.column_dimensions[get_column_letter(col)].width = 15
    
    # Hacer más ancha la columna de descripción y observación
    ws_productos.column_dimensions['B'].width = 25  # descripcion
    ws_productos.column_dimensions['K'].width = 30  # observacion
    
    # ======== HOJA DE UBICACIONES ========
    ws_ubicaciones = wb.create_sheet(title="ubicaciones")
    
    # Encabezados de ubicaciones
    headers_ubicaciones = ["id_ubicacion", "fila", "columna", "estante", "posicion", "orientacion", "deposito", "sector"]
    
    # Agregar encabezados con estilo
    for col, header in enumerate(headers_ubicaciones, 1):
        cell = ws_ubicaciones.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0F7B0F", end_color="0F7B0F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de ejemplo para ubicaciones
    ubicaciones_ejemplo = [
        [1, 1, "A", 1, 1, "NORTE", "DEP01", "PAÑOL 1"],
        [2, 1, "A", 1, 2, "SUR", "DEP01", "PAÑOL 1"],
        [3, 1, "A", 2, 1, "ESTE", "DEP01", "PAÑOL 1"],
        [4, 1, "B", 1, 1, "OESTE", "DEP01", "PAÑOL 2"],
        [5, 1, "B", 1, 2, "NORTE", "DEP01", "PAÑOL 2"],
        [6, 1, "B", 2, 1, "SUR", "DEP01", "PAÑOL 2"],
        [7, 2, "A", 1, 1, "ESTE", "DEP02", "PAÑOL 1"],
        [8, 2, "A", 1, 2, "OESTE", "DEP02", "PAÑOL 1"],
        [9, 2, "A", 2, 1, "NORTE", "DEP02", "PAÑOL 1"],
        [10, 2, "B", 1, 1, "SUR", "DEP02", "PAÑOL 2"],
        [11, 2, "B", 2, 1, "ESTE", "DEP02", "PAÑOL 2"],
        [12, 3, "A", 1, 1, "OESTE", "DEP02", "PAÑOL 2"]
    ]
    
    # Agregar datos de ubicaciones
    for row, ubicacion in enumerate(ubicaciones_ejemplo, 2):
        for col, valor in enumerate(ubicacion, 1):
            ws_ubicaciones.cell(row=row, column=col, value=valor)
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers_ubicaciones) + 1):
        ws_ubicaciones.column_dimensions[get_column_letter(col)].width = 12
    
    # ======== HOJA DE PRODUCTO_UBICACION ========
    ws_prod_ub = wb.create_sheet(title="producto_ubicacion")
    
    # Encabezados de producto_ubicacion
    headers_prod_ub = ["id_prod_ub", "codigo_producto", "id_ubicacion", "cantidad"]
    
    # Agregar encabezados con estilo
    for col, header in enumerate(headers_prod_ub, 1):
        cell = ws_prod_ub.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="D9730D", end_color="D9730D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de ejemplo para producto_ubicacion
    prod_ub_ejemplo = [
        [1, "TORN001", 1, 500],
        [2, "TURC001", 2, 300],
        [3, "CABL001", 3, 1000],
        [4, "INTE001", 4, 75],
        [5, "PINT001", 5, 25],
        [6, "RODI001", 6, 40],
        [7, "TALA001", 7, 8],
        [8, "BROC001", 8, 50],
        [9, "SOLD001", 9, 200],
        [10, "DISC001", 10, 120],
        [11, "ACEI001", 11, 80],
        [12, "FILT001", 12, 25]
    ]
    
    # Agregar datos de producto_ubicacion
    for row, prod_ub in enumerate(prod_ub_ejemplo, 2):
        for col, valor in enumerate(prod_ub, 1):
            ws_prod_ub.cell(row=row, column=col, value=valor)
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers_prod_ub) + 1):
        ws_prod_ub.column_dimensions[get_column_letter(col)].width = 15
    
    # ======== HOJA DE ALERTAS ========
    ws_alertas = wb.create_sheet(title="alertas")
    
    # Encabezados de alertas
    headers_alertas = [
        "id_alerta", "codigo_producto", "id_ubicacion", "tipo_alerta", 
        "razon", "detalles", "estado", "fecha_inicio", "fecha_fin"
    ]
    
    # Agregar encabezados con estilo
    for col, header in enumerate(headers_alertas, 1):
        cell = ws_alertas.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E03E3E", end_color="E03E3E", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de ejemplo para alertas
    alertas_ejemplo = [
        [1, "TORN001", 1, "Media", "Stock bajo", "El stock está por debajo del nivel recomendado", "sin iniciar", "2024-01-15 10:30:00", None],
        [2, "PINT001", 5, "Alta prioridad", "Producto vencido", "Revisar fecha de vencimiento del lote actual", "en progreso", "2024-01-16 14:20:00", None],
        [3, "TALA001", 7, "Baja", "Mantenimiento", "Programar mantenimiento preventivo de herramientas", "sin iniciar", "2024-01-17 09:15:00", None]
    ]
    
    # Agregar datos de alertas
    for row, alerta in enumerate(alertas_ejemplo, 2):
        for col, valor in enumerate(alerta, 1):
            ws_alertas.cell(row=row, column=col, value=valor)
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers_alertas) + 1):
        ws_alertas.column_dimensions[get_column_letter(col)].width = 15
    
    # Hacer más anchas algunas columnas específicas
    ws_alertas.column_dimensions['E'].width = 20  # razon
    ws_alertas.column_dimensions['F'].width = 30  # detalles
    
    # ======== HOJA DE INSTRUCCIONES ========
    ws_instrucciones = wb.create_sheet(title="INSTRUCCIONES")
    
    # Título
    ws_instrucciones.cell(row=1, column=1, value="INSTRUCCIONES PARA IMPORTAR DATOS")
    ws_instrucciones.cell(row=1, column=1).font = Font(bold=True, size=16, color="2383E2")
    
    instrucciones = [
        "",
        "1. ESTRUCTURA DEL ARCHIVO:",
        "   - El archivo debe contener las hojas: productos, ubicaciones, producto_ubicacion, alertas",
        "   - Los nombres de las hojas deben ser exactamente como se muestran",
        "   - La primera fila debe contener los encabezados de columna",
        "",
        "2. CAMPOS OBLIGATORIOS:",
        "   Productos: codigo, descripcion",
        "   Ubicaciones: fila, columna, estante, posicion",
        "   Producto_ubicacion: codigo_producto, id_ubicacion",
        "   Alertas: codigo_producto, tipo_alerta, razon",
        "",
        "3. TIPOS DE DATOS:",
        "   - codigo: Texto único para cada producto",
        "   - stock_actual, stock_minimo, stock_maximo: Números enteros",
        "   - fila: Números del 1 al 15, estante: del 1 al 10, posicion: Números enteros",
        "   - columna: Texto (A, B, C, etc.)",
        "   - orientacion: 'NORTE', 'SUR', 'ESTE', 'OESTE'",
        "   - deposito: 'DEP01' o 'DEP02'",
        "   - sector: 'PAÑOL 1' o 'PAÑOL 2'",
        "   - tipo_alerta: 'Alta prioridad', 'Media', 'Baja'",
        "   - estado: 'sin iniciar', 'en progreso', 'completada', 'cancelada'",
        "",
        "4. CONSEJOS:",
        "   - No modifiques los nombres de las columnas",
        "   - Asegúrate de que los códigos de producto sean únicos",
        "   - Los id_ubicacion deben corresponder a ubicaciones existentes",
        "   - Las fechas deben estar en formato YYYY-MM-DD HH:MM:SS",
        "",
        "5. IMPORTACIÓN:",
        "   - Ve a la pestaña 'Configuración' en la aplicación",
        "   - Haz clic en 'Importar desde Excel'",
        "   - Selecciona este archivo",
        "   - La aplicación validará y cargará los datos automáticamente"
    ]
    
    for row, instruccion in enumerate(instrucciones, 2):
        ws_instrucciones.cell(row=row, column=1, value=instruccion)
        if instruccion.startswith(("1.", "2.", "3.", "4.", "5.")):
            ws_instrucciones.cell(row=row, column=1).font = Font(bold=True, color="0F7B0F")
    
    # Ajustar ancho de columna
    ws_instrucciones.column_dimensions['A'].width = 80
    
    # Guardar archivo
    filename = "ejemplo_importacion_inventario.xlsx"
    wb.save(filename)
    
    print(f"✅ Archivo Excel de ejemplo creado: {filename}")
    print(f"📊 Contiene {len(productos_ejemplo)} productos de ejemplo")
    print(f"📍 Contiene {len(ubicaciones_ejemplo)} ubicaciones de ejemplo")
    print(f"⚠️ Contiene {len(alertas_ejemplo)} alertas de ejemplo")
    print("\n📋 Hojas incluidas:")
    print("   - productos: Datos de productos con stock y categorías")
    print("   - ubicaciones: Ubicaciones físicas del almacén")
    print("   - producto_ubicacion: Relación productos-ubicaciones")
    print("   - alertas: Alertas de ejemplo del sistema")
    print("   - INSTRUCCIONES: Guía completa para importar")
    
    return filename

if __name__ == "__main__":
    crear_excel_ejemplo()
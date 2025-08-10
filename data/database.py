import sqlite3
import os
from openpyxl import Workbook, load_workbook

DB_PATH = os.path.join("data", "inventario.db")

def conectar_db():
    conexion = sqlite3.connect(DB_PATH)
    conexion.execute("PRAGMA foreign_keys = ON;")
    return conexion

def crear_tablas():
    os.makedirs("data", exist_ok=True)
    conexion = conectar_db()
    cursor = conexion.cursor()

    # Tabla de productos
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS productos (
        codigo TEXT PRIMARY KEY,
        descripcion TEXT NOT NULL,
        categoria TEXT,
        proveedor TEXT,
        puesto_trabajo TEXT,
        stock_actual INTEGER DEFAULT 0,
        unidad_medida TEXT,
        tipo_control TEXT,
        stock_minimo INTEGER DEFAULT 0,
        stock_maximo INTEGER DEFAULT 0,
        observacion TEXT
    );
    """)

    # Tabla de ubicaciones
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS ubicaciones (
        id_ubicacion INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_producto TEXT NOT NULL,
        fila TEXT,
        columna TEXT,
        estante TEXT,
        deposito TEXT,
        sector TEXT,
        contenedor TEXT,
        FOREIGN KEY (codigo_producto) REFERENCES productos (codigo)
    );
    """)

    # Tabla de alertas
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS alertas (
        id_alerta INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_producto TEXT NOT NULL,
        id_ubicacion INTEGER,
        tipo_alerta TEXT,
        razon TEXT,
        detalles TEXT,
        estado TEXT DEFAULT 'sin iniciar',
        fecha_inicio TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        fecha_fin TIMESTAMP,
        FOREIGN KEY (codigo_producto) REFERENCES productos (codigo),
        FOREIGN KEY (id_ubicacion) REFERENCES ubicaciones (id_ubicacion)
    );
    """)

    conexion.commit()
    conexion.close()

def inicializar_db():
    if not os.path.exists(DB_PATH):
        print("📦 Base de datos no encontrada. Creando...")
        crear_tablas()
    else:
        print("✅ Base de datos ya existe.")

""" FUNCIONES PARA TAB_PRODUCTOS """

def buscar_productos(texto: str) -> list[tuple]:
    conn = conectar_db()
    cursor = conn.cursor()
    query = """
        SELECT p.codigo, p.descripcion, u.fila, u.columna, u.estante, u.contenedor,
               u.deposito, u.sector
        FROM productos p
        LEFT JOIN ubicaciones u ON p.codigo = u.codigo_producto
        WHERE p.codigo LIKE ?
           OR p.descripcion LIKE ?
           OR u.deposito LIKE ?
           OR u.sector LIKE ?
           OR p.puesto_trabajo LIKE ?
        ORDER BY p.codigo;
    """
    like_text = f"%{texto}%"
    cursor.execute(query, (like_text,) * 5)
    resultados = cursor.fetchall()
    conn.close()
    return resultados

def obtener_productos_con_ubicacion():
    conexion = conectar_db()
    cursor = conexion.cursor()
    cursor.execute("""
        SELECT p.codigo, p.descripcion, u.fila, u.columna, u.estante, u.contenedor,
               u.deposito, u.sector
        FROM productos p
        LEFT JOIN ubicaciones u ON p.codigo = u.codigo_producto
        ORDER BY p.codigo;
    """)
    resultado = cursor.fetchall()
    conexion.close()
    return resultado

def obtener_detalles_producto(codigo):
    conexion = conectar_db()
    cursor = conexion.cursor()
    cursor.execute("""
        SELECT proveedor, tipo_control, stock_minimo, stock_maximo,
               categoria, puesto_trabajo, observacion
        FROM productos
        WHERE codigo = ?;
    """, (codigo,))
    detalles = cursor.fetchone()
    conexion.close()
    return detalles

""" EXPORTAR / IMPORTAR EXCEL """

def exportar_a_excel(ruta_archivo):
    conn = conectar_db()
    cursor = conn.cursor()

    wb = Workbook()

    # Productos
    ws_productos = wb.active
    ws_productos.title = "productos"
    cursor.execute("SELECT * FROM productos")
    columnas = [desc[0] for desc in cursor.description]
    ws_productos.append(columnas)
    for fila in cursor.fetchall():
        ws_productos.append(fila)

    # Ubicaciones
    ws_ubicaciones = wb.create_sheet(title="ubicaciones")
    cursor.execute("SELECT * FROM ubicaciones")
    columnas = [desc[0] for desc in cursor.description]
    ws_ubicaciones.append(columnas)
    for fila in cursor.fetchall():
        ws_ubicaciones.append(fila)

    # Alertas
    ws_alertas = wb.create_sheet(title="alertas")
    cursor.execute("SELECT * FROM alertas")
    columnas = [desc[0] for desc in cursor.description]
    ws_alertas.append(columnas)
    for fila in cursor.fetchall():
        ws_alertas.append(fila)

    wb.save(ruta_archivo)
    conn.close()

def importar_desde_excel(ruta_archivo):
    conn = conectar_db()
    cursor = conn.cursor()
    wb = load_workbook(ruta_archivo)
    hojas_faltantes = []

    for nombre_hoja in ["productos", "ubicaciones", "alertas"]:
        if nombre_hoja not in wb.sheetnames:
            hojas_faltantes.append(nombre_hoja)
            continue

        ws = wb[nombre_hoja]
        filas = list(ws.iter_rows(values_only=True))
        if not filas or len(filas) < 2:
            print(f"⚠️ La hoja '{nombre_hoja}' está vacía o solo tiene encabezados.")
            continue

        columnas = filas[0]
        columnas_db = obtener_columnas_tabla(cursor, nombre_hoja)

        if list(columnas) != columnas_db:
            print(f"⚠️ Columnas de la hoja '{nombre_hoja}' no coinciden con la base de datos.")
            print(f"   Esperadas: {columnas_db}")
            print(f"   Encontradas: {list(columnas)}")
            continue

        for fila in filas[1:]:
            cursor.execute(f"""
                INSERT OR REPLACE INTO {nombre_hoja} ({','.join(columnas)})
                VALUES ({','.join(['?'] * len(columnas))})
            """, fila)

    conn.commit()
    conn.close()

    if hojas_faltantes:
        print(f"⚠️ Hojas faltantes en el Excel: {', '.join(hojas_faltantes)}")


def obtener_columnas_tabla(cursor, nombre_tabla):
    cursor.execute(f"PRAGMA table_info({nombre_tabla})")
    return [col[1] for col in cursor.fetchall()]

if __name__ == "__main__":
    inicializar_db()
